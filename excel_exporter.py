import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(all_ports, output_dir="exports"):
    """Port bilgilerini Excel dosyasına aktarır.

    Args:
        all_ports: List of dict -> {switch, port, description, status, port_type}
        output_dir: Excel dosyasının kaydedileceği klasör

    Returns:
        Oluşturulan dosyanın yolu
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Port Bilgileri"

    headers = ["Switch Adı", "Port Adı", "Description", "Status", "VLAN", "Port Tipi", "SFP Tipi", "LLDP Komşu", "MAC Adresi", "IP Adresi", "Hostname", "Hatalar", "Po Members", "Etiket"]

    # Header stilleri
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header yaz
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Veri yaz
    status_colors = {
        "connected": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "notconnect": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "disabled": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "errdisabled": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }

    type_colors = {
        "Bakır": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Fiber": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    }

    for row_idx, port in enumerate(all_ports, 2):
        error_count = port.get("error_counters", 0)
        error_str = str(error_count) if error_count > 0 else ""

        values = [
            port["switch"],
            port["port"],
            port["description"],
            port["status"],
            port.get("vlan", ""),
            port["port_type"],
            port.get("sfp_type", ""),
            port.get("lldp_neighbor", ""),
            port.get("mac_addresses", ""),
            port.get("ip_address", ""),
            port.get("hostname", ""),
            error_str,
            port.get("po_members", ""),
            port.get("tag", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # Status hücresini renklendir (D sütunu)
        status_cell = ws.cell(row=row_idx, column=4)
        status_lower = port["status"].lower() if port["status"] else ""
        if status_lower in status_colors:
            status_cell.fill = status_colors[status_lower]

        # Port tipi hücresini renklendir (F sütunu)
        type_cell = ws.cell(row=row_idx, column=6)
        if port["port_type"] in type_colors:
            type_cell.fill = type_colors[port["port_type"]]

        # Error hücresini renklendir (L sütunu)
        if error_count > 0:
            error_cell = ws.cell(row=row_idx, column=12)
            error_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            error_cell.font = Font(bold=True, color="991B1B")

    # Sütun genişliklerini ayarla
    col_widths = {"A": 20, "B": 18, "C": 40, "D": 15, "E": 10, "F": 15, "G": 16, "H": 35, "I": 30, "J": 18, "K": 30, "L": 12, "M": 25, "N": 30}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Filtre ekle
    ws.auto_filter.ref = f"A1:N{len(all_ports) + 1}"

    # Satırları dondur (header sabit)
    ws.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"port_mapping_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath
